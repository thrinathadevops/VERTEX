import io
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def generate_excel_report(drift_results: list, component_type: str, metadata: dict = None) -> io.BytesIO:
    """
    Generates a drift report matching the required Excel formatting layout based on Oracle DB format.
    """
    if metadata is None:
        metadata = {}

    # Define defaults with fallback
    current_time = datetime.now().strftime("%d-%b-%Y %H:%M:%S")
    server_name = metadata.get("Server Name", "Oracle Server")
    ip1 = metadata.get("IP Address 1", "10.9.50.24")
    ip2 = metadata.get("IP Address 2", "10.0.137.237")
    baseline = metadata.get("Baseline Name", "Default Oracle DB")
    astart1 = metadata.get("Assessment Start DateTime 1", current_time)
    astart2 = metadata.get("Assessment Start DateTime 2", current_time)
    aend1 = metadata.get("Assessment End DateTime 1", current_time)
    aend2 = metadata.get("Assessment End DateTime 2", current_time)
    instance = metadata.get("Instance (Primary)", "AXMOBILE1")

    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_empty = pd.DataFrame()
        sheet_name = "Comparison Report"
        df_empty.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Define styles
        header_fill = PatternFill(start_color="D9EAD3", fill_type="solid") # Light green header
        compliance_yes_fill = PatternFill(start_color="00B050", fill_type="solid") # Green
        compliance_no_fill = PatternFill(start_color="FF0000", fill_type="solid") # Red
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        top_left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # 1. Write Metadata Section
        metadata_rows = [
            ("Server Name", server_name),
            ("IP Address 1:", ip1),
            ("IP Address 2:", ip2),
            ("Baseline Name:", baseline),
            ("Assessment Start DateTime 1:", astart1),
            ("Assessment Start DateTime 2:", astart2),
            ("Assessment End DateTime 1:", aend1),
            ("Assessment End DateTime 2:", aend2),
            ("Instance (Primary):", instance),
            ("Compliance Score:", "") # Handled below
        ]

        # Top header "Drift {IP}"
        worksheet.merge_cells('A2:H2')
        cell_drift = worksheet.cell(row=2, column=1, value=f"_Drift {ip1}")
        cell_drift.font = bold_font
        cell_drift.alignment = center_align
        cell_drift.fill = header_fill
        for col_idx in range(1, 9):
            worksheet.cell(row=2, column=col_idx).border = thin_border

        row_idx = 3
        
        # Calculate actual compliance score dynamically
        mismatches = len([d for d in drift_results if d.get('status') == 'DRIFT'])
        total = len(drift_results)
        score = "100.00%" if total == 0 else f"{((total - mismatches) / total) * 100:.2f}%"

        for k, v in metadata_rows:
            cell_k = worksheet.cell(row=row_idx, column=1, value=k)
            cell_v = worksheet.cell(row=row_idx, column=2, value=score if "Compliance Score" in k else v)
            cell_k.border = thin_border
            cell_v.border = thin_border
            
            # Additional borders for merged look empty spaces if necessary
            for col_idx in range(3, 9):
                worksheet.cell(row=row_idx, column=col_idx).border = thin_border

            row_idx += 1

        # Leave two blank rows
        row_idx += 2

        # 2. Write Table Headers
        headers = [
            "SECTION", "SUB SECTION", "TITLE", 
            ip1, 
            ip2, 
            "Recommendedvalues", "Compliance", "Criticality"
        ]
        
        for col_idx, header_val in enumerate(headers, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=header_val)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        row_idx += 1

        # 3. Write Data Rows
        # Map drift results
        section_id = 1
        for sub_id, item in enumerate(drift_results, 1):
            parameter = item.get("parameter", "")
            val1 = str(item.get("prod_value", ""))
            val2 = str(item.get("dr_value", ""))
            status = item.get("status", "DRIFT")
            remediation = item.get("remediation", "-")
            
            is_match = (status == "MATCH")
            compliance_val = "YES" if is_match else "NO"
            
            row_data = [
                str(section_id),
                f"{section_id}.{sub_id}",
                parameter,
                val1,
                val2,
                "NA" if is_match else remediation,
                compliance_val,
                "High"
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = top_left_align
                cell.border = thin_border
                
                # Apply green/red styling for Compliance column
                if col_idx == 7: # Compliance
                    cell.alignment = center_align
                    if compliance_val == "YES":
                        cell.fill = compliance_yes_fill
                    else:
                        cell.fill = compliance_no_fill
            
            row_idx += 1

        # Adjust Column Widths to closely match the image
        column_widths = {
            'A': 12, # SECTION
            'B': 18, # SUB SECTION
            'C': 35, # TITLE
            'D': 65, # IP 1
            'E': 65, # IP 2
            'F': 30, # Recommendedvalues
            'G': 15, # Compliance
            'H': 15  # Criticality
        }
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

    output.seek(0)
    return output
